import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        if cell.value is not None:
                new_price = float(cell.value) * 0.9
                new_price_cell = sheet.cell(row, 5)
                new_price_cell.value = new_price
        else:
            print(f"Row {row} has no value in column 3")

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)
    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'f2')

    wb.save('updated_file.xlsx')


process_workbook('Python Automatic.xlsx')